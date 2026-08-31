import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Count

from apps.events.models import Event
from .models import StudentProfile, EventRegistration


@staff_member_required
def admin_analytics_dashboard(request):
    """
    Custom Django Admin view rendered with Unfold layout.
    Displays dynamic event selector, total community students, event registration count,
    top 5 attendee leaderboard, and attendee list.
    """
    events = Event.objects.all().order_by('-date')
    selected_event_id = request.GET.get('event_id')

    selected_event = None
    if selected_event_id:
        try:
            selected_event = Event.objects.get(pk=selected_event_id)
        except Event.DoesNotExist:
            selected_event = None

    if not selected_event and events.exists():
        selected_event = events.first()
        selected_event_id = str(selected_event.id)

    total_community_students = StudentProfile.objects.count()

    event_registrations_count = 0
    registered_students = []

    if selected_event:
        regs = EventRegistration.objects.filter(
            event=selected_event
        ).select_related('student', 'student__user').order_by('-registered_at')
        event_registrations_count = regs.count()
        registered_students = regs

    # Top 5 students by total event registrations
    top_students = StudentProfile.objects.annotate(
        reg_count=Count('registrations')
    ).order_by('-reg_count')[:5]

    context = {
        'title': 'Analytics & Reports',
        'events': events,
        'selected_event': selected_event,
        'selected_event_id': selected_event_id,
        'total_community_students': total_community_students,
        'event_registrations_count': event_registrations_count,
        'top_students': top_students,
        'registered_students': registered_students,
    }
    return render(request, 'admin/students/analytics_dashboard.html', context)


@staff_member_required
def admin_export_excel(request):
    """
    Generates and downloads an Excel file (.xlsx) of registered students for an event.
    """
    event_id = request.GET.get('event_id')
    if not event_id:
        return HttpResponse('event_id query parameter is required.', status=400)

    event = get_object_or_404(Event, pk=event_id)

    regs = EventRegistration.objects.filter(
        event=event
    ).select_related('student', 'student__user').order_by('-registered_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registrations"

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="00D084", end_color="00D084", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")

    headers = ["#", "Full Name", "Roll Number", "Course", "Branch", "Year", "Phone", "Section", "Registered At"]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center

    for idx, r in enumerate(regs, 1):
        sp = r.student
        row = [
            idx,
            sp.full_name or sp.user.get_full_name() or sp.user.email,
            sp.roll_number,
            sp.course,
            sp.branch,
            sp.get_academic_year_display() or sp.academic_year or '',
            sp.mobile_number,
            sp.section,
            r.registered_at.strftime("%Y-%m-%d %H:%M:%S"),
        ]
        ws.append(row)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"event_{event_id}_registrations.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
