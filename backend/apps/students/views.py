import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.db import IntegrityError
from django.db.models import Count
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.authentication import TokenAuthentication, SessionAuthentication
from rest_framework.permissions import IsAuthenticated, IsAdminUser

from apps.events.models import Event
from .models import StudentProfile, EventRegistration
from .serializers import StudentProfileSerializer, EventRegistrationSerializer


class StudentProfileView(APIView):
    """
    GET /api/student-profile/

    Returns the authenticated user's StudentProfile for autofilling the
    EventRegistrationModal. Returns an empty object ({}) instead of 404
    when the user has never registered before, so the frontend can safely
    try to autofill without error handling for first-time users.
    """
    authentication_classes = [TokenAuthentication, SessionAuthentication]
    permission_classes     = [IsAuthenticated]

    def get(self, request):
        try:
            profile = request.user.student_profile
        except StudentProfile.DoesNotExist:
            # First-time user — return empty dict so the modal opens with blank fields
            return Response({}, status=status.HTTP_200_OK)

        serializer = StudentProfileSerializer(profile)
        return Response(serializer.data)


class EventRegisterView(APIView):
    """
    POST /api/events/<event_id>/register/

    Accepts student profile details and creates/updates the StudentProfile,
    then creates the EventRegistration.
    """
    authentication_classes = [TokenAuthentication, SessionAuthentication]
    permission_classes     = [IsAuthenticated]

    def post(self, request, event_id):
        # ── 1. Fetch & validate event ─────────────────────────────────────────
        try:
            event = Event.objects.get(pk=event_id)
        except Event.DoesNotExist:
            return Response(
                {'detail': 'Event not found.'},
                status=status.HTTP_404_NOT_FOUND
            )

        # ── 2. Check registration window ──────────────────────────────────────
        if event.status.lower() == 'past' or not event.is_registration_open:
            return Response(
                {'detail': 'Registration is closed or the event has already passed.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ── 3. Validate incoming fields ───────────────────────────────────────
        serializer = EventRegistrationSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        vd = serializer.validated_data

        # ── 4. Upsert StudentProfile for this user ────────────────────────────
        profile, _ = StudentProfile.objects.update_or_create(
            user=request.user,
            defaults={
                'full_name':     vd.get('full_name', ''),
                'father_name':   vd['father_name'],
                'course':        vd['course'],
                'branch':        vd['branch'],
                'section':       vd['section'],
                'roll_number':   vd['roll_number'],
                'mobile_number': vd['mobile_number'],
                'academic_year': vd.get('academic_year') or None,
            }
        )

        # ── 5. Create EventRegistration (duplicate → IntegrityError) ──────────
        try:
            EventRegistration.objects.create(event=event, student=profile)
        except IntegrityError:
            return Response(
                {'detail': 'You have already registered for this event.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        return Response(
            {'detail': 'Registration successful.'},
            status=status.HTTP_201_CREATED
        )


class AnalyticsView(APIView):
    """
    GET /api/students/analytics/?event_id=X

    Protected: Only Admin / Staff users (IsAdminUser) can access metrics and attendee lists.
    """
    authentication_classes = [TokenAuthentication, SessionAuthentication]
    permission_classes     = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        event_id = request.query_params.get('event_id')

        total_community_students = StudentProfile.objects.count()

        # Top 5 students by registration count
        top_students_qs = StudentProfile.objects.annotate(
            reg_count=Count('registrations')
        ).order_by('-reg_count')[:5]

        top_students = [
            {
                'id': s.id,
                'name': s.full_name or s.user.get_full_name() or s.user.email,
                'branch': s.branch,
                'academic_year': s.get_academic_year_display() or s.academic_year,
                'count': s.reg_count,
                'mobile_number': s.mobile_number,
            }
            for s in top_students_qs
        ]

        event_registrations_count = 0
        registered_list = []

        if event_id:
            try:
                regs = EventRegistration.objects.filter(
                    event_id=event_id
                ).select_related('student', 'student__user').order_by('-registered_at')

                event_registrations_count = regs.count()

                for r in regs:
                    sp = r.student
                    registered_list.append({
                        'id': r.id,
                        'name': sp.full_name or sp.user.get_full_name() or sp.user.email,
                        'roll_number': sp.roll_number,
                        'course': sp.course,
                        'branch': sp.branch,
                        'academic_year': sp.get_academic_year_display() or sp.academic_year,
                        'section': sp.section,
                        'mobile_number': sp.mobile_number,
                        'registered_at': r.registered_at.isoformat(),
                    })
            except Exception as e:
                print(f"Error fetching event registrations for analytics: {e}")
        else:
            event_registrations_count = EventRegistration.objects.count()

        return Response({
            'total_community_students': total_community_students,
            'event_registrations': event_registrations_count,
            'top_students': top_students,
            'registered_list': registered_list,
        })


class ExportExcelView(APIView):
    """
    GET /api/students/analytics/export-excel/?event_id=X

    Generates and downloads an Excel spreadsheet of students registered for the event.
    Protected: Strictly IsAdminUser (IsAuthenticated + IsAdminUser).
    """
    authentication_classes = [TokenAuthentication, SessionAuthentication]
    permission_classes     = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        event_id = request.query_params.get('event_id')
        if not event_id:
            return Response({'detail': 'event_id query parameter is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            event = Event.objects.get(pk=event_id)
        except Event.DoesNotExist:
            return Response({'detail': 'Event not found.'}, status=status.HTTP_404_NOT_FOUND)

        regs = EventRegistration.objects.filter(
            event=event
        ).select_related('student', 'student__user').order_by('-registered_at')

        # Create Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registrations"

        # Excel Styles
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="00D084", end_color="00D084", fill_type="solid")
        alignment_center = Alignment(horizontal="center", vertical="center")

        headers = ["#", "Full Name", "Roll Number", "Course", "Branch", "Year", "Phone", "Father Name", "Section", "Registered At"]
        ws.append(headers)

        # Style header row
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center

        # Append data rows
        for idx, r in enumerate(regs, 1):
            sp = r.student
            row = [
                idx,
                sp.full_name or sp.user.get_full_name() or sp.user.email,
                sp.roll_number,
                sp.course,
                sp.branch,
                sp.get_academic_year_display() or sp.academic_year or '',
                sp.mobile_number,
                sp.father_name,
                sp.section,
                r.registered_at.strftime("%Y-%m-%d %H:%M:%S"),
            ]
            ws.append(row)

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Stream response
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"event_{event_id}_registrations.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
